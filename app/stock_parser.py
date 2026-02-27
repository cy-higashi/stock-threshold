# -*- coding: utf-8 -*-
"""
日次在庫数ファイル（CSV/TSV/TXT/XLSX）のパースと商品コード別在庫合算。
日次在庫数ディレクトリを再帰的に検索し、該当ファイルを取得する。
"""
import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# 対象拡張子（日次在庫数ファイル）
DATA_EXTENSIONS = (".csv", ".tsv", ".txt", ".xlsx")

# エンコーディングのフォールバック順（テキスト用。BOM 付き UTF-8 を先に試す）
ENCODINGS = ("utf-8-sig", "utf-8", "cp932")

# Choice ポータル: TSV ジョイン用のカラム位置のデフォルト（0-based）
CHOICE_DETAILS_PRODUCT_CODE_COL_DEFAULT = 102  # 103列目
CHOICE_CHANGE_STOCK_COL_DEFAULT = 3  # 4列目


def _detect_delimiter(sample: str) -> str:
    """先頭行から区切り文字を推定。"""
    if "\t" in sample and sample.count("\t") >= sample.count(","):
        return "\t"
    return ","


def _read_csv_rows(path: Path, has_header: bool, product_column: str, stock_column: str) -> list[tuple[str, int]]:
    """CSV/TSV/TXT をパースし、(商品コード, 在庫数) のリストを返す。"""
    pairs: list[tuple[str, int]] = []
    for enc in ENCODINGS:
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                sample = f.readline()
                f.seek(0)
                delimiter = _detect_delimiter(sample)
                if has_header:
                    reader = csv.DictReader(f, delimiter=delimiter)
                    rows = list(reader)
                else:
                    reader = csv.reader(f, delimiter=delimiter)
                    rows = [{str(i): c for i, c in enumerate(row)} for row in reader]
                for row in rows:
                    # BOM 付きヘッダーのフォールバック（\ufeff が先頭につくことがある）
                    code = row.get(product_column) or row.get("\ufeff" + product_column)
                    stock_raw = row.get(stock_column) or row.get("\ufeff" + stock_column)
                    if code is None or code == "" or stock_raw is None:
                        continue
                    try:
                        stock = int(float(str(stock_raw).replace(",", "").strip()))
                    except (ValueError, TypeError):
                        continue
                    key = str(code).strip()
                    if key:
                        pairs.append((key, stock))
                return pairs
        except (UnicodeDecodeError, csv.Error, OSError):
            continue
    return pairs


def _read_xlsx_rows(path: Path, has_header: bool, product_column: str, stock_column: str) -> list[tuple[str, int]]:
    """XLSX をパースし、(商品コード, 在庫数) のリストを返す。"""
    pairs: list[tuple[str, int]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return pairs
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        if has_header:
            try:
                code_idx = header.index(product_column)
            except ValueError:
                code_idx = 0
            try:
                stock_idx = header.index(stock_column) if stock_column else 1
            except ValueError:
                stock_idx = 1 if code_idx == 0 else 0
        else:
            code_idx = int(product_column) if product_column.isdigit() else 0
            stock_idx = int(stock_column) if stock_column.isdigit() else 1
        data_rows = rows[1:] if has_header else rows
        for row in data_rows:
            if row is None or len(row) <= max(code_idx, stock_idx):
                continue
            code = row[code_idx]
            stock_raw = row[stock_idx]
            if code is None or str(code).strip() == "":
                continue
            try:
                stock = int(float(str(stock_raw or "0").replace(",", "").strip()))
            except (ValueError, TypeError):
                continue
            pairs.append((str(code).strip(), stock))
    finally:
        wb.close()
    return pairs


def _parse_choice_tsv_join(daily_stock_dir: Path, portal_config: dict[str, Any]) -> dict[str, int]:
    """
    Choice ポータル専用: 2つのTSVを第一カラムでジョインし、返礼品コードと在庫数を取得する。
    - 末尾が _change_stock のTSV: mapping.change_stock_column_index 列目を在庫数
    - 末尾が _change_stock でないTSV: mapping.details_product_code_column_index 列目を返礼品コード
    stg_ はステージングのため無視し、末尾 _change_stock で判別する。
    """
    mapping = portal_config.get("mapping") or {}
    details_col = mapping.get("details_product_code_column_index", CHOICE_DETAILS_PRODUCT_CODE_COL_DEFAULT)
    stock_col = mapping.get("change_stock_column_index", CHOICE_CHANGE_STOCK_COL_DEFAULT)
    change_stock_suffix = "_change_stock"
    tsv_files = list(daily_stock_dir.rglob("*.tsv"))
    change_stock_paths: list[Path] = []
    details_paths: dict[str, Path] = {}  # base_without_suffix -> path
    for p in tsv_files:
        if not p.is_file():
            continue
        stem = p.stem
        if stem.endswith(change_stock_suffix):
            change_stock_paths.append(p)
        else:
            details_paths[stem] = p

    aggregated: dict[str, int] = {}
    for change_stock_path in change_stock_paths:
        stem = change_stock_path.stem
        details_base = stem[: -len(change_stock_suffix)]
        details_path = details_paths.get(details_base)
        if not details_path or not details_path.exists():
            continue

        change_stock_rows = _read_tsv_rows(change_stock_path, stock_col, is_stock=True)
        details_rows = _read_tsv_rows(details_path, details_col, is_stock=False)
        if not change_stock_rows or not details_rows:
            continue

        details_by_key = {k: v for k, v in details_rows}
        for join_key, stock in change_stock_rows:
            product_code = details_by_key.get(join_key)
            if product_code and product_code.strip():
                aggregated[product_code.strip()] = aggregated.get(product_code.strip(), 0) + stock
    return aggregated


def _read_tsv_rows(path: Path, target_col: int, is_stock: bool) -> list[tuple[str, str | int]]:
    """TSV を読み、先頭列をキー、target_col 列を値としたリストを返す。ヘッダーなし。"""
    result: list[tuple[str, str | int]] = []
    for enc in ENCODINGS:
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                reader = csv.reader(f, delimiter="\t")
                for row in reader:
                    if len(row) <= max(0, target_col):
                        continue
                    key = str(row[0]).strip()
                    val = row[target_col]
                    if not key:
                        continue
                    if is_stock:
                        try:
                            stock = int(float(str(val or "0").replace(",", "").strip()))
                            result.append((key, stock))
                        except (ValueError, TypeError):
                            continue
                    else:
                        result.append((key, str(val).strip()))
                return result
        except (UnicodeDecodeError, csv.Error, OSError):
            continue
    return result


def parse_portal_stock(
    daily_stock_dir: Path,
    portal_config: dict[str, Any],
) -> dict[str, int]:
    """
    日次在庫数ディレクトリを再帰的に検索し、CSV/TSV/TXT/XLSX をパースする。
    商品コードで在庫数を合算した辞書を返す。
    Choice ポータルは tsv_join_mode で TSV 2ファイルのジョイン処理を行う。
    カラム名は portal_config.mapping で指定（product_code_column, stock_column 等）。
    """
    if portal_config.get("tsv_join_mode"):
        return _parse_choice_tsv_join(daily_stock_dir, portal_config)

    mapping = portal_config.get("mapping") or {}
    has_header = mapping.get("has_header", portal_config.get("has_header", True))
    if has_header:
        product_column = mapping.get("product_code_column") or "商品コード"
        stock_column = mapping.get("stock_column") or "在庫数"
    else:
        product_column = str(mapping.get("product_code_column_index", 0))
        stock_column = str(mapping.get("stock_column_index", 1))

    aggregated: dict[str, int] = {}

    if not daily_stock_dir.is_dir():
        return aggregated

    for path in daily_stock_dir.rglob("*"):
        if not path.is_file():
            continue
        suf = path.suffix.lower()
        if suf not in DATA_EXTENSIONS:
            continue

        if suf == ".xlsx":
            pairs = _read_xlsx_rows(path, has_header, product_column, stock_column)
        else:
            pairs = _read_csv_rows(path, has_header, product_column, stock_column)

        for code, stock in pairs:
            if code:
                aggregated[code] = aggregated.get(code, 0) + stock

    return aggregated
