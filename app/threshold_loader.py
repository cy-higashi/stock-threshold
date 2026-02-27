# -*- coding: utf-8 -*-
"""
最低在庫数定義ファイル（CSV/Excel）から返礼品コードと最低在庫数を読み込む。
各ポータルの min_stock_base_path で指定した CSV または XLSX ファイルのパスを参照する。
"""
import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# 返礼品コード列の候補（CSV/Excel のヘッダー名）
PRODUCT_CODE_HEADERS = ("返礼品コード", "商品コード", "出品者SKU", "管理コード")
MIN_STOCK_HEADER = "最低在庫数"

ENCODINGS = ("utf-8", "utf-8-sig", "cp932")


def _load_from_csv(path: Path) -> dict[str, int]:
    """CSV から返礼品コード・最低在庫数を読み込む。"""
    result: dict[str, int] = {}
    for enc in ENCODINGS:
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                sample = f.readline()
                delimiter = "\t" if "\t" in sample and sample.count("\t") >= sample.count(",") else ","
                f.seek(0)
                reader = csv.DictReader(f, delimiter=delimiter)
                headers = [h.strip() for h in (reader.fieldnames or [])]
                code_col = None
                min_col = None
                for h in PRODUCT_CODE_HEADERS:
                    if h in headers:
                        code_col = h
                        break
                if MIN_STOCK_HEADER in headers:
                    min_col = MIN_STOCK_HEADER
                if not code_col or not min_col:
                    return result
                for row in reader:
                    code = (row.get(code_col) or "").strip()
                    min_val = row.get(min_col)
                    if not code:
                        continue
                    try:
                        min_stock = int(float(str(min_val or "0").replace(",", "").strip()))
                    except (ValueError, TypeError):
                        continue
                    result[code] = min_stock
                return result
        except (UnicodeDecodeError, csv.Error, OSError):
            continue
    return result


def _load_from_xlsx(path: Path, portal_config: dict[str, Any]) -> dict[str, int]:
    """Excel から返礼品コード・最低在庫数を読み込む。"""
    mapping = portal_config.get("mapping") or {}
    product_column = mapping.get("product_code_column") or "返礼品コード"
    result: dict[str, int] = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return result
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        code_idx = 0
        for h in (product_column,) + PRODUCT_CODE_HEADERS:
            try:
                code_idx = header.index(h)
                break
            except ValueError:
                continue
        try:
            min_idx = header.index(MIN_STOCK_HEADER)
        except ValueError:
            min_idx = 1 if code_idx == 0 else 0
        for row in rows[1:]:
            if row is None or len(row) <= max(code_idx, min_idx):
                continue
            code = row[code_idx]
            min_val = row[min_idx]
            if code is None or str(code).strip() == "":
                continue
            try:
                min_stock = int(float(str(min_val).replace(",", "").strip()))
            except (ValueError, TypeError):
                continue
            result[str(code).strip()] = min_stock
    finally:
        wb.close()
    return result


def load_thresholds(portal_min_stock_path: str, portal_config: dict[str, Any]) -> dict[str, int]:
    """
    ポータル用の最低在庫数定義ファイルを読み込む。
    portal_min_stock_path は CSV または XLSX ファイルへの直接パス。
    setting.json の portals.{ポータル名}.min_stock_base_path で指定する。
    返礼品コードをキー・最低在庫数を値とした辞書を返す。
    """
    path = Path(portal_min_stock_path)
    if not path.exists():
        return {}
    suf = path.suffix.lower()
    if suf == ".csv":
        return _load_from_csv(path)
    if suf in (".xlsx", ".xls"):
        return _load_from_xlsx(path, portal_config)
    return {}
